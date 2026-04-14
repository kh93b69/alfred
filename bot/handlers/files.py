import os
import logging
import json
import zipfile
from aiogram import Router
from aiogram.types import Message

from bot.services.knowledge import KNOWLEDGE_DIR

router = Router()
logger = logging.getLogger(__name__)

# Поддерживаемые форматы
SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx", ".xls", ".csv", ".xmind"}


def extract_text_from_pdf(filepath: str) -> str:
    """Извлекает текст из PDF"""
    from PyPDF2 import PdfReader
    reader = PdfReader(filepath)
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def extract_text_from_docx(filepath: str) -> str:
    """Извлекает текст из DOCX"""
    from docx import Document
    doc = Document(filepath)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def extract_text_from_xlsx(filepath: str) -> str:
    """Извлекает текст из Excel таблицы"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        lines.append(f"### Лист: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            # Фильтруем пустые ячейки
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def extract_text_from_csv(filepath: str) -> str:
    """Извлекает текст из CSV"""
    with open(filepath, "r", encoding="utf-8") as f:
        return f.read()


def extract_text_from_xmind(filepath: str) -> str:
    """Извлекает текст из XMind карты"""
    # XMind файл — это ZIP-архив с content.json внутри
    lines = []
    try:
        with zipfile.ZipFile(filepath, "r") as z:
            # XMind 8+ хранит данные в content.json
            if "content.json" in z.namelist():
                with z.open("content.json") as f:
                    data = json.loads(f.read())
                    _parse_xmind_topic(data, lines, level=0)
            # Старый формат — content.xml
            elif "content.xml" in z.namelist():
                lines.append("[XMind XML формат — извлечён как текст]")
                with z.open("content.xml") as f:
                    lines.append(f.read().decode("utf-8"))
    except Exception as e:
        lines.append(f"[Ошибка чтения XMind: {e}]")
    return "\n".join(lines)


def _parse_xmind_topic(data, lines: list, level: int):
    """Рекурсивно парсит структуру XMind карты"""
    if isinstance(data, list):
        for item in data:
            _parse_xmind_topic(item, lines, level)
        return

    if isinstance(data, dict):
        # Корневая тема
        root = data.get("rootTopic", data)
        title = root.get("title", "")
        if title:
            prefix = "  " * level + ("- " if level > 0 else "# ")
            lines.append(f"{prefix}{title}")

        # Дочерние темы
        children = root.get("children", {})
        attached = children.get("attached", [])
        for child in attached:
            _parse_xmind_topic(child, lines, level + 1)


def extract_text(filepath: str, extension: str) -> str:
    """Извлекает текст из файла в зависимости от формата"""
    if extension == ".pdf":
        return extract_text_from_pdf(filepath)
    elif extension == ".docx":
        return extract_text_from_docx(filepath)
    elif extension in (".xlsx", ".xls"):
        return extract_text_from_xlsx(filepath)
    elif extension == ".csv":
        return extract_text_from_csv(filepath)
    elif extension == ".xmind":
        return extract_text_from_xmind(filepath)
    else:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()


@router.message(lambda m: m.document is not None)
async def handle_document(message: Message):
    """Обработчик загрузки файлов — сохраняет в базу знаний"""
    doc = message.document
    filename = doc.file_name or "без_имени.txt"
    _, ext = os.path.splitext(filename.lower())

    # Проверяем формат
    if ext not in SUPPORTED_EXTENSIONS:
        await message.answer(
            f"Формат {ext} пока не поддерживается.\n"
            f"Поддерживаемые: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )
        return

    await message.answer(f"Обрабатываю файл: {filename}...")

    try:
        # Скачиваем файл во временную папку
        temp_path = os.path.join(KNOWLEDGE_DIR, f"_temp_{filename}")
        file = await message.bot.get_file(doc.file_id)
        await message.bot.download_file(file.file_path, temp_path)

        # Извлекаем текст
        text = extract_text(temp_path, ext)

        # Удаляем временный файл
        os.remove(temp_path)

        if not text.strip():
            await message.answer("Файл пустой или не удалось извлечь текст.")
            return

        # Сохраняем как .txt в базу знаний
        safe_name = filename.rsplit(".", 1)[0]
        save_path = os.path.join(KNOWLEDGE_DIR, f"{safe_name}.txt")
        with open(save_path, "w", encoding="utf-8") as f:
            f.write(text)

        # Считаем размер
        words = len(text.split())
        await message.answer(
            f"Файл сохранён в базу знаний: **{safe_name}**\n"
            f"Слов: ~{words}\n\n"
            f"Теперь я буду учитывать эту информацию при ответах.",
            parse_mode="Markdown",
        )

    except Exception as e:
        logger.error(f"Ошибка обработки файла {filename}: {e}")
        # Убираем временный файл если остался
        if os.path.exists(temp_path):
            os.remove(temp_path)
        await message.answer(f"Ошибка при обработке файла: {e}")
